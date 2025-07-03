import os
import json
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


def generateTxt(name, code, onError):
    try:
        # 读取 JSON 文件
        with open(os.path.join("temp", code + ".json"), "r", encoding="utf-8") as f:
            jsonData = json.load(f)

        # 当前时间与路径设置
        time = pd.Timestamp.now()
        formatTime = time.strftime('%Y-%m-%d')
        dir_path = os.path.join("result", formatTime)
        os.makedirs(dir_path, exist_ok=True)

        txt_path = os.path.join(dir_path, f"{name}({code}) {formatTime}.txt")

        with open(txt_path, "w", encoding="utf-8") as f_txt:

            # 写入【股票信息】
            if "股票信息" in jsonData:
                f_txt.write("【股票信息】\n")
                for key, value in jsonData["股票信息"].items():
                    f_txt.write(f"{key}: {value}\n")
                f_txt.write("\n")

            # 写入【今日概览】（A股才有）
            if "今日概览" in jsonData:
                f_txt.write("【今日概览】\n")
                for key, value in jsonData["今日概览"].items():
                    f_txt.write(f"{key}: {value}\n")
                f_txt.write("\n")

            # 写入【数据指标】
            if "数据指标" in jsonData:
                f_txt.write("【数据指标】\n")
                for key, value in jsonData["数据指标"].items():
                    f_txt.write(f"{key}: {value}\n")

        print(f"✅ TXT 文件已保存到: {txt_path}")

    except Exception as e:
        # 出错日志
        with open("error.log", "a", encoding="utf-8") as f_log:
            f_log.write(f"{code} - {str(e)}\n")
        onError(f"{code} 生成 TXT 失败")


def generateExcel(name, code, onError):
    try:
        # 读取 JSON 文件
        with open(os.path.join("temp", code + ".json"), "r", encoding="utf-8") as f:
            jsonData = json.load(f)

        # 创建文件保存目录
        time = pd.Timestamp.now()
        formatTime = time.strftime('%Y-%m-%d')
        dir_path = os.path.join("result", formatTime)
        os.makedirs(dir_path, exist_ok=True)

        excel_path = os.path.join(dir_path, f"{name}({code}) {formatTime}.xlsx")

        # 创建 Excel 工作簿
        wb = Workbook()

        def write_dict_to_sheet(sheet_name, data_dict):
            ws = wb.create_sheet(title=sheet_name)
            for row_idx, (key, value) in enumerate(data_dict.items(), start=1):
                ws.cell(row=row_idx, column=1, value=key)
                ws.cell(row=row_idx, column=2, value=value)
                # 自动列宽（可选）
                ws.column_dimensions[get_column_letter(1)].width = 20
                ws.column_dimensions[get_column_letter(2)].width = 40

        # 写入【股票信息】
        if "股票信息" in jsonData:
            write_dict_to_sheet("股票信息", jsonData["股票信息"])

        # 写入【今日概览】（A 股特有）
        if "今日概览" in jsonData:
            write_dict_to_sheet("今日概览", jsonData["今日概览"])

        # 写入【数据指标】
        if "数据指标" in jsonData:
            write_dict_to_sheet("数据指标", jsonData["数据指标"])

        # 删除默认的空 sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # 保存文件
        wb.save(excel_path)
        print(f"✅ Excel 文件已保存到: {excel_path}")

    except Exception as e:
        with open("error.log", "a", encoding="utf-8") as f_log:
            f_log.write(f"{code} - {str(e)}\n")
        onError(f"{code} 生成 Excel 失败")